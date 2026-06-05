import json
import random
from datetime import datetime, timedelta, timezone
from urllib.parse import quote

from django.conf import settings
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.files.storage import FileSystemStorage
from django.http import (
    Http404,
    HttpResponse,
    HttpResponseForbidden,
    HttpResponseNotAllowed,
    HttpResponsePermanentRedirect,
    HttpResponseRedirect,
    JsonResponse,
)
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.utils.decorators import method_decorator
from django.utils.safestring import mark_safe
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook

from apps.race.pricing import create_team_payment, upsert_team_extras
from website.forms import NewsPostForm, PageForm, TeamForm
from website.models import (
    Checkpoint,
    CheckpointTag,
    Race,
    RaceAdmin,
    TakenKP,
    Team,
    TeamMemberRaceLog,
    TeamStartLog,
)
from website.models.race import Category, RegStatus
from website.sync_xlsx import import_file_xlsx

from ..models.news import MenuItem, Page


def is_admin(user):
    return user.is_superuser


def is_race_admin(user, race):
    if not user.is_authenticated:
        return False
    return RaceAdmin.objects.filter(race=race, user=user).exists()


class AddNewsPostView(View):
    def get(self, request, race_slug):
        return HttpResponseNotAllowed(["POST"])

    def post(self, request, race_slug):
        if not request.user.is_authenticated:
            return HttpResponseRedirect(
                reverse("login") + "?next=" + quote(request.path, safe="/:@")
            )
        race = get_object_or_404(Race, slug=race_slug)
        if not is_race_admin(request.user, race):
            return HttpResponseForbidden()
        form = NewsPostForm(request.POST, request.FILES)
        if form.is_valid():
            post = form.save(commit=False)
            post.race = race
            post.save()
            return HttpResponseRedirect(
                reverse("race", kwargs={"race_slug": race_slug})
            )
        from apps.race.views import RacePageView

        context = RacePageView.build_context(race, request.user)
        context["post_form"] = form
        return render(request, "race/race_page.html", context)


class RaceIdRedirectView(View):
    def get(self, request, race_id, **kwargs):
        race = get_object_or_404(Race, pk=race_id)
        new_path = request.path.replace(f"/race/{race_id}/", f"/race/{race.slug}/", 1)
        qs = f"?{request.GET.urlencode()}" if request.GET else ""
        return HttpResponsePermanentRedirect(new_path + qs)


@method_decorator(user_passes_test(is_admin, login_url="login"), name="dispatch")
class TeamMemberRaceLogView(View):
    template_name = "website/team_member_race_logs.html"

    def get(self, request, race_slug):
        race = get_object_or_404(Race, slug=race_slug)
        logs = (
            TeamMemberRaceLog.objects.select_related("member_tag")
            .filter(race=race)
            .order_by("-finish_time", "-start_time", "member_tag__number")
        )

        start_logs = TeamStartLog.objects.select_related("team").filter(
            race=race, team__isnull=False
        )

        tag_to_team = {}
        tag_to_start_timestamp = {}
        for start_log in start_logs:
            member_tags = start_log.member_tags or []
            for value in member_tags:
                value = (value or "").strip()
                if not value:
                    continue
                normalized = value.upper()
                if normalized not in tag_to_team:
                    tag_to_team[normalized] = start_log.team
                    tag_to_start_timestamp[normalized] = start_log.start_timestamp

        def format_timestamp(timestamp_ms: int) -> str:
            if not timestamp_ms:
                return ""
            try:
                dt = datetime.fromtimestamp(timestamp_ms / 1000, tz=timezone.utc)
            except (TypeError, ValueError, OSError, OverflowError):
                return str(timestamp_ms)
            return dt.astimezone().strftime("%d.%m.%Y %H:%M:%S")

        entries = []
        for log in logs:
            tag = log.member_tag
            tag_uid = (tag.tag_id or "").strip()
            normalized = tag_uid.upper()
            team = tag_to_team.get(normalized)
            fallback_start = tag_to_start_timestamp.get(normalized, 0)
            effective_start = log.start_time or fallback_start
            entries.append(
                {
                    "log": log,
                    "tag_number": tag.number,
                    "tag_uid": tag_uid,
                    "team": team,
                    "team_label": self._make_team_label(team),
                    "start_timestamp": effective_start,
                    "raw_start_timestamp": log.start_time,
                    "finish_timestamp": log.finish_time,
                    "start_display": format_timestamp(effective_start),
                    "finish_display": format_timestamp(log.finish_time),
                }
            )

        return render(
            request,
            self.template_name,
            {
                "race": race,
                "entries": entries,
            },
        )

    @staticmethod
    def _make_team_label(team):
        if not team:
            return ""
        name = team.teamname or "Без названия"
        start_number = team.start_number or "—"
        return f"{name} (#{start_number})"


def index_dummy(request):
    return render(request, "website/index_dummy.html")


def new_point(request, pk):
    if request.method == "POST":
        points = request.POST.get("points")
        for point in points.split(","):
            TakenKP.objects.create(
                team_id=pk,
                point_number=point.lstrip().rstrip(),
                image_url="manual",
                status="new",
                timestamp=int(datetime.now().timestamp()) * 1000,
                year=2024,
            )
    return HttpResponseRedirect("/race/1/teams_result")


def update_protocol(request):
    if not request.user.is_staff:
        raise Http404("File not found.")

    wb = load_workbook(filename=settings.PROTOCOL_DIR + "protokol.xlsx")
    # grab the active worksheet
    sheet_tabs = {
        "6h": "6ч",
        "12h_ww": "12ч_ЖЖ",
        "12h_mw": "12ч_МЖ",
        "12h_mm": "12ч_ММ",
        "12h_team": "12ч_группа",
        "24h": "24ч",
    }
    distance_max_time = {
        "6h": timedelta(hours=6),
        "12h_ww": timedelta(hours=12),
        "12h_mw": timedelta(hours=12),
        "12h_mm": timedelta(hours=12),
        "12h_team": timedelta(hours=12),
        "24h": timedelta(hours=24),
    }

    for tab in sheet_tabs:
        ws = wb[sheet_tabs[tab]]
        # export KP
        cpoints = Checkpoint.objects.filter(year="2024").order_by("iterator")
        column = 14
        point_column = {}
        for p in cpoints:
            ws.cell(8, column, p.number)
            ws.cell(9, column, p.cost)
            point_column[p.number] = column
            column += 1
        # export teams
        Team().update_places()
        teams = Team.objects.filter(category=tab, year="2024").order_by(
            "place", "start_number"
        )
        teams = [team for team in teams if team.paid_sum > 0]
        line = 10
        for team in teams:
            row = str(line)
            ws["B" + row] = team.start_number
            ws["C" + row] = team.teamname
            athlets = [
                team.athlet1,
                team.athlet2,
                team.athlet3,
                team.athlet4,
                team.athlet5,
                team.athlet6,
            ]
            athlets = ", ".join(athlets[: int(team.paid_people)])
            ws["D" + row] = athlets
            ws["E" + row] = (
                team.start_time + timedelta(hours=5) if team.start_time else ""
            )
            ws["F" + row] = (
                team.finish_time + timedelta(hours=5) if team.finish_time else ""
            )
            if team.distance_time:
                ws["G" + row] = team.distance_time
                if team.distance_time > distance_max_time[tab]:
                    ws["H" + row] = team.distance_time - distance_max_time[tab]
                else:
                    ws["H" + row] = ""
            points = TakenKP.objects.filter(team=team)
            points_sum = 0
            points_count = 0
            for point in points:
                points_sum += point.point.cost
                points_count += 1
                ws.cell(line, point_column[point.point.number], 1)
            ws["I" + row] = points_count
            ws["J" + row] = team.points_sum + team.penalty
            ws["K" + row] = team.penalty
            ws["L" + row] = team.points_sum
            ws["M" + row] = team.place if team.place != 10000 else 0
            if team.dnf:
                ws["M" + row] = "СН"
                ws["K" + row] = "снятие"
            line += 1

    # Save the file
    filename = "protokol2024.xlsx"
    wb.save(settings.PROTOCOL_DIR + filename)
    return render(
        request,
        "website/save_protokol.html",
        {"success": "save", "file_url": settings.PROTOCOL_URL + filename},
    )


def upload_protocol(request):
    if not request.user.is_staff:
        raise Http404("File not found.")

    if request.method == "POST" and request.FILES["myfile"]:
        myfile = request.FILES["myfile"]
        fs = FileSystemStorage()

        curr_time = datetime.now(timezone.utc) + timedelta(hours=5)
        file_prefix = (
            "uploads/"
            + str(curr_time.year)
            + str(curr_time.month).zfill(2)
            + str(curr_time.day).zfill(2)
            + "_"
            + str(curr_time.hour).zfill(2)
            + str(curr_time.minute).zfill(2)
            + str(curr_time.second).zfill(2)
            + "_"
        )

        filename = fs.save(settings.PROTOCOL_DIR + file_prefix + myfile.name, myfile)

        uploaded_file_url = fs.url(settings.PROTOCOL_URL + file_prefix + myfile.name)

        # read this file:
        err, msg = import_file_xlsx(filename)

        return render(
            request,
            "website/simple_upload.html",
            {"uploaded_file_url": uploaded_file_url, "err": err, "msg": msg},
        )
    return render(request, "website/simple_upload.html")


def regulations(request):
    return render(request, "website/regulations.html")


def privacy_policy(request):
    """Display privacy policy page."""
    return HttpResponseRedirect(reverse("page", args=["privacy_app"]))


def refund_policy(request):
    """Display refund policy page."""
    return HttpResponseRedirect(reverse("page", args=["refund_policy"]))


def service_order_rules(request):
    """Display service order rules page."""
    return HttpResponseRedirect(reverse("page", args=["service_order_rules"]))


def rules(request):
    """Display service order rules page."""
    return HttpResponseRedirect(reverse("page", args=["rules"]))


def contacts(request):
    """Display contacts page."""
    return HttpResponseRedirect(reverse("page", args=["contacts"]))


def page(request, slug):
    """Display a static page based on the slug."""
    try:
        page = Page.objects.get(slug=slug)
    except Page.DoesNotExist:
        raise Http404("Page not found.")

    context = {
        "page": page,
        "menu": MenuItem.objects.all(),
        "is_moderator": request.user.is_authenticated
        and request.user.groups.filter(name="Moderators").exists(),
    }
    return render(request, "website/static_page.html", context=context)


@login_required
def edit_page(request, slug):
    """Allow moderators to edit a static page."""
    try:
        page = Page.objects.get(slug=slug)
    except Page.DoesNotExist:
        raise Http404("Page not found.")

    if not request.user.groups.filter(name="Moderators").exists():
        raise Http404("Page not found.")

    if request.method == "POST":
        form = PageForm(request.POST, instance=page)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse("page", args=[page.slug]))
    else:
        form = PageForm(instance=page)

    context = {
        "form": form,
        "page": page,
        "menu": MenuItem.objects.all(),
        "is_moderator": True,
    }
    return render(request, "website/edit_page.html", context=context)


# API _________________________________________________________________
class RaceView(View):
    def get(self, request):
        races = Race.objects.filter(is_active=True)
        data = []
        for race in races:
            data.append(
                {
                    "id": race.id,
                    "name": race.name,
                    "date": race.date.strftime("%Y-%m-%d"),
                    "is_active": race.is_active,
                }
            )
        return JsonResponse(data, safe=False)


@method_decorator(csrf_exempt, name="dispatch")
class TeamsTimesView(View):
    def post(self, request):
        try:
            times = json.loads(request.body)

            for time_ in times:
                try:
                    team = Team.objects.get(id=time_.get("team_id"))
                except Team.DoesNotExist:
                    continue

                fields = []
                if time_.get("start_time") and not team.start_time:
                    team.start_time = time_.get("start_time")
                    fields.append("start_time")

                if time_.get("finish_time") and not team.finish_time:
                    team.finish_time = time_.get("finish_time")
                    fields.append("finish_time")

                if fields:
                    team.save(update_fields=fields)

            return JsonResponse(
                {"message": "Teams times updated."},
                status=200,
            )

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data."}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)


@method_decorator(csrf_exempt, name="dispatch")
class PointTagsView(View):
    def post(self, request, race_id):
        try:
            data = json.loads(request.body)

            point_number = data.get("point_number")
            point = self.get_point_by_number(point_number)
            if not point:
                return JsonResponse(
                    {"error": f"Point with number {point_number} not found."},
                    status=404,
                )

            tag_id = data.get("tag_id")
            if tag_id is None:
                return JsonResponse(
                    {"error": "tag_id is a required field."}, status=400
                )

            _, created = CheckpointTag.objects.update_or_create(
                point=point, tag_id=tag_id
            )
            if created:
                return JsonResponse(
                    {"message": "PointTag created successfully."}, status=201
                )
            return JsonResponse(
                {"message": f"PointTag with tag_id {tag_id} updated."},
                status=200,
            )

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data."}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    def get_point_by_number(self, point_number):
        try:
            return Checkpoint.objects.get(number=point_number, year=2024)
        except Checkpoint.DoesNotExist:
            return None


def teams_api(request):
    """Возвращает список команд"""
    query_params = request.GET.get("category", "")
    teams = (
        Team.objects.filter(year=2024, paid_people__gt=0)
        .order_by("id")
        .values(
            "id",
            "paid_people",
            "dist",
            "category",
            "teamname",
            "start_number",
            "teamname",
            "city",
            "organization",
            "year",
            "start_number",
            "start_time",
            "finish_time",
            "place",
            "dnf",
            "penalty",
            "athlet1",
            "athlet2",
            "athlet3",
            "athlet4",
            "athlet5",
            "athlet6",
        )
    )

    if query_params:
        teams = teams.filter(category=query_params)

    teams = list(teams)
    for team in teams:
        paid_people = team.get("paid_people")
        members = [team["athlet1"]]
        if paid_people > 1 and team["athlet2"]:
            members.append(team["athlet2"])

        if paid_people > 2 and team["athlet3"]:
            members.append(team["athlet3"])

        if paid_people > 3 and team["athlet4"]:
            members.append(team["athlet4"])

        if paid_people > 4 and team["athlet5"]:
            members.append(team["athlet5"])

        if paid_people > 5 and team["athlet6"]:
            members.append(team["athlet6"])
        for i in range(6):
            key = f"athlet{i+1}"
            team.pop(key)
        teamname_description = ", ".join(members)
        team["teamname"] = team["teamname"] + f" ({teamname_description})"

    return JsonResponse(list(teams), safe=False)


@csrf_exempt
def upload_photo(request, race_id):
    """save file from post request"""
    if request.method != "POST":
        raise Http404("File not found.")

    race = Race.objects.filter(id=race_id).first()
    if not race:
        return JsonResponse({"error": "race not found"}, status=404)

    if not race.is_photo_upload_enabled:
        return JsonResponse({"error": "photo upload is disabled"}, status=403)

    file = request.FILES.get("photo")
    team_id = request.POST["team_id"]
    point_number = request.POST["point_number"]
    timestamp = request.POST.get("timestamp")
    nfc = request.POST.get("nfc")
    phone_uuid = request.POST.get("phone_uuid")

    team = Team.objects.filter(id=team_id).first()
    if not team:
        return JsonResponse({"error": "team not found"}, status=404)

    uploaded_file_url = ""
    if file:
        start_number = team.start_number
        folder_name = (
            f"photos/{race_id}/"
            + f"{start_number}-{team_id}/{phone_uuid}/{point_number}-{file.name}"
        )
        fs = FileSystemStorage()
        filename = fs.save(folder_name, file)
        uploaded_file_url = fs.url(filename)

    # save to db
    TakenKP.objects.create(
        team=team,
        point_number=point_number,
        image_url=uploaded_file_url,
        timestamp=timestamp,
        nfc=nfc,
        phone_uuid=phone_uuid,
    )
    return JsonResponse({"success": True})


def build_category_options(race_id, current_category_id=None, team=None):
    """Category options carrying ``data-counts`` expanded from min/max people.

    If current_category_id names an inactive category, it is prepended so that
    existing teams whose category was later deactivated still render correctly.

    Each option also carries ``remaining`` — free slots for that category
    (``""`` when the category is unlimited), self-excluding ``team`` from the
    occupancy so a team editing within its own category isn't counted against
    its own limit. The client uses it to disable full categories.
    """
    cats = list(Category.objects.filter(race_id=race_id, is_active=True))
    if current_category_id is not None:
        active_ids = {c.id for c in cats}
        if current_category_id not in active_ids:
            current_cat = Category.objects.filter(
                id=current_category_id, race_id=race_id
            ).first()
            if current_cat:
                cats.insert(0, current_cat)
    options = []
    for category in cats:
        remaining = category.remaining_people(exclude_team=team)
        options.append(
            {
                "id": category.id,
                "label": f"{category.short_name} ({category.name})",
                "counts": list(range(category.min_people, category.max_people + 1)),
                "remaining": "" if remaining is None else int(remaining),
                "is_current": category.id == current_category_id,
            }
        )
    return options


def build_team_form_context(race, team, is_edit=False, bypass_limits=False, form=None):
    """Unified context shared by the add and edit team forms.

    The ``extras`` list (config island) mirrors ``compute_team_charge`` —
    see ``apps/race/pricing.py``. Keep the per-extra ``price``/``freePerTeam``/
    ``count``/``countPaid`` fields and the live-total formula in sync with
    ``src/static/js/team-form.js``.
    """
    current_category_id = getattr(team, "category2_id", None)
    price_tiers = race.price_tier_ladder()
    current_price = race.current_price
    race_remaining = race.remaining_people(exclude_team=team)

    team_extra_map = {}
    if team and team.pk:
        team_extra_map = {te.race_extra_id: te for te in team.extras.all()}
    extras_config = []
    for extra in race.extras.filter(is_active=True):
        te = team_extra_map.get(extra.id)
        count = te.count if te else 0
        # Preserve a submitted/invalid POST value across re-render when bound.
        if form is not None:
            field = f"extra_{extra.code}"
            if field in form.fields:
                try:
                    count = int(form[field].value() or 0)
                except (TypeError, ValueError):
                    count = te.count if te else 0
        extras_config.append(
            {
                "code": extra.code,
                "name": extra.name,
                "price": extra.price,
                "freePerTeam": extra.free_per_team,
                "count": count,
                "countPaid": te.count_paid if te else 0,
            }
        )

    config = {
        "currentPrice": current_price,
        "paidPeople": team.paid_people,
        "extras": extras_config,
        "isEdit": is_edit,
        "raceRemaining": race_remaining,
        "currentCategoryId": current_category_id,
        "bypassLimits": bypass_limits,
    }
    return {
        "current_price": current_price,
        "paid_people": team.paid_people,
        "price_tiers": price_tiers,
        "reg_open": race.reg_status == RegStatus.OPEN,
        "is_editable": race.is_teams_editable,
        "reg_status": race.reg_status,
        "extras": extras_config,
        "category_options": build_category_options(
            race.id, current_category_id, team=team
        ),
        "team_form_config_json": mark_safe(json.dumps(config)),
    }


class AddTeam(View):
    def get(self, request, race_slug):
        race = get_object_or_404(Race, slug=race_slug)

        if not request.user.is_authenticated:
            return HttpResponseRedirect(
                reverse("account_start") + f"?next={request.path}"
            )
        form = TeamForm(race.id)
        return render(
            request,
            "website/add_team.html",
            {
                "race": race,
                "race_id": race.id,
                "team_form": form,
                "team": Team(),
                "action": reverse("add_team", args=[race.slug]),
                **build_team_form_context(
                    race, Team(), bypass_limits=request.user.is_superuser
                ),
            },
        )

    def post(self, request, race_slug):
        if not request.user.is_authenticated:
            return HttpResponseRedirect(
                reverse("account_start") + f"?next={request.path}"
            )

        race = get_object_or_404(Race, slug=race_slug)
        if not race.is_teams_editable:
            return HttpResponse("Регистрация закрыта")

        category2_id = request.POST.get("category2_id")
        category2 = Category.objects.filter(id=category2_id).first()
        if not category2:
            return JsonResponse({"error": "Category not found"}, status=404)

        payment_method = request.GET.get("method", "sbp2")
        if payment_method != "sbp2":
            raise Http404

        data = request.POST.copy()
        data["dist"] = category2.code
        data["paymentid"] = "%016x" % random.randrange(16**16)  # legacy
        form = TeamForm(race.id, data, bypass_limits=request.user.is_superuser)
        if form.is_valid():
            # save team (extra_<code> fields are add-ons, not Team columns)
            team_fields = {
                k: v for k, v in form.cleaned_data.items() if not k.startswith("extra_")
            }
            team: Team = Team.objects.create(
                year=race.date.year,
                owner_id=request.user.id,
                **team_fields,
            )

            upsert_team_extras(team, form.cleaned_data, race)

            if race.reg_status != RegStatus.OPEN:
                return HttpResponseRedirect(reverse("my_teams", args=[race.slug]))

            # payment (race fee + add-on deltas, one VTB/SBP order)
            response = create_team_payment(request, team, race)
            if response is None:
                return HttpResponseRedirect(reverse("my_teams", args=[race.slug]))
            return response

        return render(
            request,
            "website/add_team.html",
            {
                "race": race,
                "race_id": race.id,
                "team_form": form,
                "team": Team(),
                "action": reverse("add_team", args=[race.slug]),
                **build_team_form_context(
                    race, Team(), bypass_limits=request.user.is_superuser, form=form
                ),
            },
        )


class AllTeamsResultView(View):
    def get(self, request, race_slug, category_id=None):
        try:
            race = Race.objects.get(slug=race_slug)
        except Race.DoesNotExist:
            # page not found
            raise Http404(f"Гонка {race_slug} не найдена.")

        race_id = race.id
        teams_ = (
            Team.objects.filter(category2__race_id=race_id, paid_people__gt=0)
            .exclude(start_time=0)
            .select_related("category2")
        )

        if category_id:
            teams_ = teams_.filter(category2_id=category_id)

        points = Checkpoint.objects.filter(race_id=race_id, cost__gte=0)
        cost = {}
        for p in points:
            cost[p.number] = p.cost

        for team in teams_:
            # members
            members = [team.athlet1]
            if team.ucount > 1 and team.athlet2:
                members.append(team.athlet2)

            if team.ucount > 2 and team.athlet3:
                members.append(team.athlet3)
            if team.ucount > 3 and team.athlet4:
                members.append(team.athlet4)
            if team.ucount > 4 and team.athlet5:
                members.append(team.athlet5)
            if team.ucount > 5 and team.athlet6:
                members.append(team.athlet6)
            team.members = ", ".join(members)

            team.points_nfc = []
            summ_nfc = 0
            count_nfc = 0
            nfc_points = (
                TakenKP.objects.filter(team=team.id)
                .exclude(nfc="")
                .order_by("timestamp")
            )
            nfc_points_set = set()
            for point in nfc_points:
                if (
                    cost[point.point_number]
                    and point.point_number not in nfc_points_set
                ):
                    summ_nfc += cost[point.point_number]
                    count_nfc += 1
                    team.points_nfc.append(point.point_number)
                    nfc_points_set.add(point.point_number)
            team.summ_nfc = summ_nfc
            team.count_nfc = count_nfc

            # ищем уникальные чипы нфс
            unique_nfc_tags = set()
            for point in nfc_points:
                nfc_tags = point.nfc.split(",")
                unique_nfc_tags.update(nfc_tags)

            team.unique_nfc_tags_count = len(unique_nfc_tags)

            # PHOTO
            team.points_photo = []
            summ_photo = 0
            count_photo = 0
            photo_points = (
                TakenKP.objects.filter(team=team.id)
                .exclude(image_url="")
                .distinct("point_number")
                .order_by("point_number")
            )
            for p in photo_points:
                if cost[p.point_number] and p.point_number not in team.points_nfc:
                    summ_photo += cost[p.point_number]
                    count_photo += 1
                    team.points_photo.append(p.point_number)

            team.summ_photo = summ_photo
            team.count_photo = count_photo

            team.summ_both = team.summ_photo + team.summ_nfc

            team.time_diff = team.finish_time - team.start_time
            if team.finish_time == 0:
                team.time_diff = 0
            seconds = int(team.time_diff / 1000)
            minutes = int(seconds / 60)
            hour = int(minutes / 60)
            if team.time_diff:
                team.time = f"{hour}:{minutes%60:02d}:{seconds%60:02d}"
            else:
                team.time = "-"

            if team.category2_id == 16 and minutes > 6 * 60:
                team.penalty = minutes - 6 * 60
            if team.category2_id in (17, 18, 19, 20) and minutes > 12 * 60:
                team.penalty = minutes - 12 * 60
            if team.category2_id == 21 and minutes > 24 * 60:
                team.penalty = minutes - 24 * 60
            if team.category2_id in (22, 23) and minutes > 8 * 60:
                team.penalty = minutes - 8 * 60

            team.points_nfc = ", ".join(str(p) for p in team.points_nfc)
            team.points_photo = ", ".join(str(p) for p in team.points_photo)

            team.summ_after_penalty = team.summ_both - team.penalty
            if team.category == "6h":
                team.category = "06h"

        teams_ = sorted(
            teams_,
            key=lambda x: (x.category, -x.summ_after_penalty, x.time_diff),
        )
        category = ""
        counter = 1
        for team in teams_:
            if category != team.category:
                counter = 1
                category = team.category
            team.place = counter
            counter += 1

        category = None
        if category_id:
            category = Category.objects.filter(id=category_id).first()

        context = {
            "race": race,
            "teams": teams_,
            "show_category": True,
            "category": category,
        }
        return render(request, "teams_result.html", context)


class TeamPointsView(View):
    def get(self, request, team_id):
        team = Team.objects.filter(id=team_id).first()
        photo_points = (
            TakenKP.objects.filter(team=team_id, timestamp__gt=1697223600000)
            .exclude(image_url="")
            .distinct("point_number", "image_url")
            .order_by("point_number")
        )
        seconds = int((team.finish_time - team.start_time) / 1000)
        minutes = int(seconds / 60)
        hour = int(minutes / 60)
        time = (f"{hour}:{minutes%60:02d}:{seconds%60:02d}",)

        context = {"team": team, "photo_points": photo_points, "time": time}
        return render(request, "team_points.html", context)
