from datetime import datetime, timedelta

from openpyxl import load_workbook

from website.models import Checkpoint, TakenKP, Team


def import_file_xlsx(filename):
    wb = load_workbook(filename=filename, data_only=True)
    if "start" in wb.get_sheet_names():
        err, msg = import_file_start(wb["start"])
        if err:
            return err, msg
    if "finish" in wb.get_sheet_names():
        err, msg = import_file_start(wb["finish"], finish=True)
        if err:
            return err, msg
    if "points" in wb.get_sheet_names():
        err, msg = import_points(wb["points"])
        if err:
            return err, msg
    if "Финиш" in wb.get_sheet_names():
        err, msg = import_vova_points(wb["Финиш"])
        if err:
            return err, msg
    return False, msg


def import_file_start(ws, finish=False):
    row = 3
    team_number = ws.cell(row, 1).value
    teams = set()
    while team_number:
        if not Team.objects.filter(start_number=team_number, year="2019").exists():
            return True, "Команда с номером %s не найдена [%s,%s]" % (
                team_number,
                row,
                1,
            )
        teams.add(team_number)
        row += 1
        team_number = ws.cell(row, 1).value
    teams_count = row - 3

    start_datetime = {}
    for i in range(teams_count):
        row = i + 3
        team_number = int(ws.cell(row, 1).value)
        start_date = str(ws.cell(row, 2).value)
        start_time = str(ws.cell(row, 3).value)
        if start_date and len(start_date) != 8:
            return True, "Неправильная дата '%s' в ячейке [%s,%s]" % (
                start_date,
                row,
                2,
            )
        if start_date and not start_date.isdigit():
            return True, "Неправильная дата '%s' в ячейке [%s,%s]" % (
                start_date,
                row,
                2,
            )

        if start_time and len(start_time) != 6:
            return True, "Неправильное время '%s' в ячейке [%s,%s]" % (
                start_time,
                row,
                3,
            )
        if start_time and not start_time.isdigit():
            return True, "Неправильное время '%s' в ячейке [%s,%s]" % (
                start_time,
                row,
                3,
            )

        if start_time and start_date:
            try:
                d = datetime(
                    int(start_date[4:8]),
                    int(start_date[2:4]),
                    int(start_date[:2]),
                    int(start_time[:2]),
                    int(start_time[2:4]),
                    int(start_time[4:6]),
                )
                start_datetime[team_number] = d - timedelta(hours=5)
            except Exception:
                return True, "Неправильное время и дата '%s' '%s' в строке %s" % (
                    start_date,
                    start_time,
                    row,
                )
    # all good, import it
    count_updated = 0
    for team_number in start_datetime:
        if Team.objects.filter(start_number=team_number, year="2019").count() > 1:
            return True, "Команд с номером %s больше 2х" % team_number
        team = Team.objects.filter(start_number=team_number, year="2019").get()
        time_is_same = True
        if finish:
            t_old = team.finish_time
        else:
            t_old = team.start_time
        t_new = start_datetime[team_number]
        if not t_old:
            time_is_same = False
        if t_old:
            if t_old.year != t_new.year:
                time_is_same = False
            if t_old.month != t_new.month:
                time_is_same = False
            if t_old.day != t_new.day:
                time_is_same = False
            if t_old.hour != t_new.hour:
                time_is_same = False
            if t_old.minute != t_new.minute:
                time_is_same = False
            if t_old.second != t_new.second:
                time_is_same = False
        if not time_is_same:
            if finish:
                team.finish_time = t_new
            else:
                team.start_time = t_new
            count_updated += 1
            team.save()
    msg = "Обновлено %s записей (прочитано %s в файле)" % (count_updated, teams_count)
    return False, msg


def import_points(ws):
    row = 3
    team_number = ws.cell(row, 1).value
    teams = set()
    while team_number:
        team_number = int(team_number)
        if not Team.objects.filter(start_number=team_number, year="2019").exists():
            return True, "Команда с номером %s не найдена [%s,%s]" % (
                team_number,
                row,
                1,
            )
        if team_number in teams:
            return True, "Команда номером %s встречается дважды [%s,%s]" % (
                team_number,
                row,
                1,
            )
        teams.add(team_number)
        row += 1
        team_number = ws.cell(row, 1).value
    teams_count = row - 3

    import_point_count = 0
    import_point_success = 0

    all_points = {}
    for i in range(teams_count):
        row = i + 3
        points_count = int(ws.cell(row, 2).value)
        team_number = ws.cell(row, 1).value
        if team_number:
            team_number = int(team_number)

        all_points[team_number] = set()
        for p in range(points_count):
            column = p + 3
            point = ws.cell(row, column).value
            if point:
                point = int(point)
                if not Checkpoint.objects.filter(number=point, year="2019").exists():
                    return True, "КП с номером %s не найден [%s,%s]" % (
                        point,
                        row,
                        column,
                    )
                if point in all_points[team_number]:
                    return True, "КП с номером %s встречается дважды [%s,%s]" % (
                        point,
                        row,
                        column,
                    )
                all_points[team_number].add(point)
    for team_start_num in all_points:
        team = Team.objects.filter(start_number=team_start_num).get()
        for point_num in all_points[team_start_num]:
            import_point_count += 1
            point = Checkpoint.objects.filter(number=str(point_num)).get()
            if not TakenKP.objects.filter(team=team, point=point).exists():
                new_point = TakenKP(team=team, point=point)
                new_point.save()
                import_point_success += 1

    msg = "Обновлено %s точек (в файле найдено %s)" % (
        import_point_success,
        import_point_count,
    )

    return False, msg


def import_vova_points(ws):
    line = 9
    team_num = int(ws.cell(line, 2).value)
    msg = ""

    import_point_success = 0
    while team_num:
        print("team_num", team_num)
        team = Team.objects.filter(start_number=str(team_num)).get()
        for p in range(7, 56):
            point_num = int(ws.cell(8, p).value)
            print(point_num)
            is_point_taken = ws.cell(line, p).value

            if is_point_taken:
                print("point_num", point_num)
                point = Checkpoint.objects.filter(number=str(point_num)).get()
                if not TakenKP.objects.filter(team=team, point=point).exists():
                    new_point = TakenKP(team=team, point=point)
                    new_point.save()
                    import_point_success += 1
        line += 1
        team_num = int(ws.cell(line, 2).value)

    msg = "Обновлено %s точек" % (import_point_success)

    return False, msg
