import datetime
import os

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from donate.models import ClubMember, DonationPeriod, MemberDonation

DEFAULT_FILE = os.path.join(
    os.path.dirname(__file__),
    "../..",
    "Взносы ГШ_Тамила.xlsx",
)

YEAR_SHEETS = ["2022", "2023", "2024", "2025", "2026"]


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ("%d.%m.%Y", "%d.%m.%y"):
            try:
                return datetime.datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def parse_recipient(value):
    if value is None:
        return ""
    s = str(value).strip()
    if "Тамила" in s:
        return MemberDonation.RECIPIENT_TAMILA
    return ""


class Command(BaseCommand):
    help = "One-time import of dues from Взносы ГШ_Тамила.xlsx"

    def add_arguments(self, parser):
        parser.add_argument(
            "file_path",
            nargs="?",
            default=DEFAULT_FILE,
            help="Path to the Excel file",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Print what would be imported without touching the DB",
        )

    def handle(self, *args, **options):
        file_path = os.path.abspath(options["file_path"])
        dry_run = options["dry_run"]

        if not os.path.exists(file_path):
            raise CommandError(f"File not found: {file_path}")

        self.stdout.write(f"Loading {file_path}")
        wb = openpyxl.load_workbook(file_path, data_only=True)

        members_created = 0
        periods_created = 0
        donations_created = 0
        donations_updated = 0

        for sheet_name in YEAR_SHEETS:
            if sheet_name not in wb.sheetnames:
                self.stdout.write(
                    self.style.WARNING(f"Sheet {sheet_name} not found, skipping")
                )
                continue

            year = int(sheet_name)
            ws = wb[sheet_name]

            # Get or create the two periods for this year
            period_i_name = f"Весна {year}"
            period_ii_name = f"Осень {year}"

            if not dry_run:
                period_i, created = DonationPeriod.objects.get_or_create(
                    name=period_i_name,
                    defaults={"date": datetime.date(year, 1, 1), "is_active": True},
                )
                if created:
                    periods_created += 1

                period_ii, created = DonationPeriod.objects.get_or_create(
                    name=period_ii_name,
                    defaults={"date": datetime.date(year, 7, 1), "is_active": True},
                )
                if created:
                    periods_created += 1
            else:
                period_i = period_ii = None

            row_count = 0
            for row in ws.iter_rows(min_row=4, values_only=True):
                name_raw = row[0]
                if name_raw is None:
                    continue
                name = str(name_raw).strip()
                if not name:
                    continue

                # col indices: C=2, D=3, E=4, F=5, G=6, H=7
                amount_i = row[2]
                date_i = parse_date(row[3])
                amount_ii = row[4]
                date_ii = parse_date(row[5])
                recipient_raw = row[6]
                note_raw = row[7]
                note = str(note_raw).strip() if note_raw is not None else ""
                recipient = parse_recipient(recipient_raw)

                if not dry_run:
                    member, created = ClubMember.objects.get_or_create(name=name)
                    if created:
                        members_created += 1
                else:
                    member = None

                # First half
                if amount_i is not None or date_i is not None:
                    row_count += 1
                    if dry_run:
                        self.stdout.write(
                            f"  [DRY] {period_i_name} | {name} | "
                            f"amount={amount_i} date={date_i} "
                            f"recipient={recipient} note={note!r}"
                        )
                    else:
                        donation, created = MemberDonation.objects.get_or_create(
                            member=member, period=period_i
                        )
                        donation.is_paid = amount_i is not None
                        donation.amount = amount_i
                        donation.paid_date = date_i
                        donation.recipient = recipient
                        donation.note = note
                        donation.save()
                        if created:
                            donations_created += 1
                        else:
                            donations_updated += 1

                # Second half
                if amount_ii is not None or date_ii is not None:
                    row_count += 1
                    if dry_run:
                        self.stdout.write(
                            f"  [DRY] {period_ii_name} | {name} | "
                            f"amount={amount_ii} date={date_ii} "
                            f"recipient={recipient} note={note!r}"
                        )
                    else:
                        donation, created = MemberDonation.objects.get_or_create(
                            member=member, period=period_ii
                        )
                        donation.is_paid = amount_ii is not None
                        donation.amount = amount_ii
                        donation.paid_date = date_ii
                        donation.recipient = recipient
                        donation.note = note
                        donation.save()
                        if created:
                            donations_created += 1
                        else:
                            donations_updated += 1

            self.stdout.write(
                f"  Sheet {sheet_name}: {row_count} donation records processed"
            )

        if dry_run:
            self.stdout.write(self.style.WARNING("Dry run — no changes written to DB"))
        else:
            self.stdout.write(
                self.style.SUCCESS(
                    f"Done. Members created: {members_created}, "
                    f"Periods created: {periods_created}, "
                    f"Donations created: {donations_created}, "
                    f"Donations updated: {donations_updated}"
                )
            )
