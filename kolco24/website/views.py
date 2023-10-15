import json
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from time import gmtime, strftime, time

from django.conf import settings
from django.contrib.auth import login as auth_login
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.files.storage import FileSystemStorage
from django.db.models import Count, OuterRef, Subquery, Sum
from django.http import Http404, HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook
from website.email import send_login_email
from website.forms import (
    Export2GoogleDocsForm,
    FastLoginForm,
    LoginForm,
    RegForm,
    TeamForm,
    TeamFormAdmin,
)
from website.googledocs import (
    export_payments_to_sheet,
    export_teams,
    export_teams_pretty,
    import_category_from_sheet,
    import_start_numbers_from_sheet,
)
from website.models import (
    Athlet,
    ControlPoint,
    FastLogin,
    Payment,
    PaymentLog,
    PaymentsYa,
    PointTag,
    Race,
    TakenKP,
    Team,
)
from website.models.race import Category
from website.sync_xlsx import import_file_xlsx


def is_admin(user):
    return user.is_superuser


def index(request):
    init_val = {}
    myteams = []
    free_athlets = 0
    if request.user.is_authenticated:
        init_val = {
            "first_name": request.user.first_name,
            "last_name": request.user.last_name,
            "email": request.user.email,
            "phone": request.user.profile.phone,
        }
        myteams = Team.objects.filter(owner=request.user, year=2023)
        free_athlets = Athlet.objects.filter(owner=request.user, team=None).count()
    reg_form = RegForm(request.POST or None, initial=init_val)
    reg_form.set_user(request.user)

    if request.method == "POST" and reg_form.is_valid():
        user = reg_form.reg_user()
        auth_login(request, user)
        return HttpResponseRedirect("/team")

    teams_count, members_count = Team.get_info()

    contex = {
        "cost": PaymentsYa.get_cost(),
        "reg_form": reg_form,
        "team_count": teams_count,
        "people_count": int(members_count),
        "myteams": myteams,
        "myteams_count": len(myteams),
        "free_athlet": free_athlets,
        "reg_open": settings.REG_OPEN,
    }
    return render(request, "website/index.html", contex)


def index_dummy(request):
    return render(request, "website/index_dummy.html")


def passlogin(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect("/")
    form = LoginForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        user = form.authenticate_user()
        auth_login(request, user)
        return HttpResponseRedirect("/")

    return render(request, "website/passlogin.html", {"form": form})


def login(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect("/")
    form = FastLoginForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        login = FastLogin()
        login_key = login.new_login_link(form.cleaned_data["email"])
        send_login_email(form.cleaned_data["email"], login_key)
        return render(request, "website/login.html", {"success": "ok", "form": form})

    return render(request, "website/login.html", {"form": form})


def login_by_key(request, login_key=""):
    keys = FastLogin.objects.filter(login_key=login_key)
    for key in keys:
        curr_time = datetime.now(timezone.utc)
        delta = curr_time - key.created_at
        if delta.total_seconds() < 24 * 3600:
            if (
                request.method == "POST"
                and "login_key" in request.POST
                and request.POST["login_key"] == login_key
            ):
                user = key.user
                auth_login(request, user)
                return HttpResponseRedirect("/")
            else:
                return render(
                    request,
                    "website/login.html",
                    {
                        "success": "enter",
                        "login_key": login_key,
                        "username": key.user.first_name + " " + key.user.last_name,
                    },
                )
    return HttpResponseRedirect("/login")


def logout_user(request):
    """Logout user"""
    if request.method == "POST" and request.POST.get("logout", "") == "logout":
        if request.user.is_authenticated:
            logout(request)
    return HttpResponseRedirect("/")


def teams(request, template=""):
    teams = [
        {
            "teams": Team.objects.filter(
                dist="6h", year="2023", category="", paid_people__gt=0
            ).order_by("start_number", "id"),
            "dist_name": "6ч",
        },
        {
            "teams": Team.objects.filter(
                dist="12h", year="2023", category="", paid_people__gt=0
            ).order_by("start_number", "id"),
            "dist_name": "12ч",
        },
        {
            "teams": Team.objects.filter(
                dist="24h", year="2023", category="", paid_people__gt=0
            ).order_by("start_number", "id"),
            "dist_name": "25ч",
        },
        {
            "teams": Team.objects.filter(
                category="24h", year="2023", paid_people__gt=0
            ).order_by("start_number", "id"),
            "dist_name": '"Хождение по мукам" (24ч, 4-6 человек)',
        },
        {
            "teams": Team.objects.filter(
                category="12h_team", year="2023", paid_people__gt=0
            ).order_by("start_number", "id"),
            "dist_name": '"Горе от ума" (12ч, 4-6 человек)',
        },
        {
            "teams": Team.objects.filter(
                category="12h_mw", year="2023", paid_people__gt=0
            ).order_by("start_number", "id"),
            "dist_name": '"Горе от ума" (12ч, МЖ)',
        },
        {
            "teams": Team.objects.filter(
                category="12h_ww", year="2023", paid_people__gt=0
            ).order_by("start_number", "id"),
            "dist_name": '"Горе от ума" (12ч, ЖЖ)',
        },
        {
            "teams": Team.objects.filter(
                category="12h_mm", year="2023", paid_people__gt=0
            ).order_by("start_number", "id"),
            "dist_name": '"Горе от ума" (12ч, ММ)',
        },
        {
            "teams": Team.objects.filter(
                category="6h", year="2023", paid_people__gt=0
            ).order_by("start_number", "id"),
            "dist_name": '"Денискины рассказы" (6ч, 2-3 человека)',
        },
        # {
        #     'teams': Team.objects.filter(year='2023').order_by('start_number'),
        #     'dist_name': 'Кольцо 2023'
        # },
        # {
        #     'teams': Team.objects.filter(year='10').order_by('start_number'),
        #     'dist_name': 'Отказ от участия'
        # },
    ]

    if request.user.is_superuser:
        teams.append(
            {
                "teams": Team.objects.filter(paid_people=0, year=2023),
                "dist_name": "Неоплаченное",
            }
        )

    context = {
        "teams": teams,
    }
    return render(request, "website/teams%s.html" % template, context)


def teams_predstart(request):
    return teams(request, template="_predstart")


def teams_start(request):
    teams = [
        {
            "teams": Team.objects.filter(
                paid_sum__gt=1, category="6h", start_time__isnull=True, year=2023
            ),
            "dist_name": "Дистанция 6ч",
        },
        {
            "teams": Team.objects.filter(
                paid_sum__gt=1,
                category__startswith="12h",
                start_time__isnull=True,
                year=2023,
            ),
            "dist_name": "Дистанция 12ч",
        },
        {
            "teams": Team.objects.filter(
                paid_sum__gt=1, category="24h", start_time__isnull=True, year=2023
            ),
            "dist_name": "Дистанция 24ч",
        },
        {
            "teams": Team.objects.filter(
                paid_sum__gt=1, start_time__isnull=False, year=2023
            ).order_by("start_time"),
            "dist_name": "Стартовавшие",
        },
    ]
    for teamgroup in teams:
        for team in teamgroup["teams"]:
            if team.start_time:
                team.start_time = team.start_time + timedelta(hours=5)
            if team.finish_time:
                team.finish_time = team.finish_time + timedelta(hours=5)

    context = {
        "teams": teams,
    }
    return render(request, "website/teams_start.html", context)


def teams_finish(request):
    teams = [
        {
            "teams": Team.objects.filter(
                paid_sum__gt=1, category="6h", finish_time__isnull=True, year=2023
            ).order_by("start_number"),
            "dist_name": "Дистанция 6ч",
        },
        {
            "teams": Team.objects.filter(
                paid_sum__gt=1,
                category__startswith="12h",
                finish_time__isnull=True,
                year=2023,
            ).order_by("start_number"),
            "dist_name": "Дистанция 12ч",
        },
        {
            "teams": Team.objects.filter(
                paid_sum__gt=1, category="24h", finish_time__isnull=True, year=2023
            ).order_by("start_number"),
            "dist_name": "Дистанция 24ч",
        },
        {
            "teams": Team.objects.filter(
                paid_sum__gt=1, finish_time__isnull=False, year=2023
            ).order_by("finish_time"),
            "dist_name": "Финишировавшие",
        },
    ]
    for teamgroup in teams:
        for team in teamgroup["teams"]:
            if team.start_time:
                team.start_time = team.start_time + timedelta(hours=5)
            if team.finish_time:
                team.finish_time = team.finish_time + timedelta(hours=5)

    context = {
        "teams": teams,
    }
    return render(request, "website/teams_finish.html", context)


def teams_protocol(request):
    teams = [
        {
            "teams": Team.objects.filter(paid_sum__gt=1, category="6h", year=2023),
            "dist_name": "Дистанция 6ч",
        },
        {
            "teams": Team.objects.filter(
                paid_sum__gt=1, category__startswith="12h", year=2023
            ),
            "dist_name": "Дистанция 12ч",
        },
        {
            "teams": Team.objects.filter(paid_sum__gt=1, category="24h", year=2023),
            "dist_name": "Дистанция 24ч",
        },
    ]
    for teamgroup in teams:
        for team in teamgroup["teams"]:
            if team.start_time:
                team.start_time = team.start_time + timedelta(hours=5)

    context = {
        "teams": teams,
    }
    return render(request, "website/teams_protocol.html", context)


def success(request, teamid=""):
    team = Team.objects.filter(paymentid=teamid, year=2023)[:1]
    if team:
        context = {
            "team": team[0],
        }
        return render(request, "website/success.html", context)
    raise Http404("File not found.")


@login_required
def my_team(request, teamid="", template="my_team"):
    team_form = TeamForm(request.POST or None)
    paymentid = team_form.init_vals(request.user, teamid)
    if not paymentid:
        return HttpResponseRedirect("/teams")

    # if request.user.is_superuser:
    #     team_form_admin = TeamFormAdmin(None)
    #     team_form_admin.init_vals(request.user, teamid)

    cost_now = PaymentsYa.get_cost()

    if request.method == "GET":
        if teamid != paymentid:
            return HttpResponseRedirect("/team/%s" % paymentid)

        main_team = Team.objects.get(paymentid=paymentid, year=2023)
        other_teams = Team.objects.filter(owner=request.user, year=2023).exclude(
            paymentid=paymentid
        )
        teams_count, _ = Team.get_info()
        # история платежей
        payments = Payment.objects.filter(team=main_team, status="done")
        # проверка бухалтерии
        payments_sum = 0
        for p in payments:
            payments_sum += p.additional_charge + p.payment_amount
        payments_ok = payments_sum == main_team.paid_sum

        context = {
            "cost": cost_now,
            "team_form": team_form,
            "other_teams": other_teams,
            "main_team": main_team,
            "curr_time": datetime.now(timezone.utc) + timedelta(hours=5),
            "timestamp": time(),
            "reg_open": settings.REG_OPEN,
            "additional_charge": main_team.additional_charge,
            "payments": payments,
            "payments_ok": payments_ok,
            "a": payments_sum,
            "b": main_team.paid_sum,
            "payback_sum": main_team.paid_sum * 0.92,
        }
        # if request.user.is_superuser:
        #     context["team_form_admin"] = team_form_admin
        return render(request, "website/%s.html" % template, context)
    elif request.method == "POST" and team_form.is_valid():
        if team_form.access_possible(request.user):
            team = team_form.save()
            if not team:
                Http404("Wrong values")
            response_data = {}
            response_data["sum"] = (team.ucount - team.paid_people) * cost_now

            new_event = PaymentLog(
                team=team, payment_method="save", paid_sum=response_data["sum"]
            )
            new_event.save()
            response_data["success"] = "true"
            return JsonResponse(response_data)
    raise Http404("Wrong values")


def team_predstart(request, teamid=""):
    if request.user.is_superuser:
        return my_team(request, teamid, "team_predstart")
    raise Http404("Not found")


def team_start(request, teamid=""):
    if request.user.is_superuser:
        return my_team(request, teamid, "team_start")
    raise Http404("Not found")


def team_finish(request, teamid=""):
    if request.user.is_superuser:
        return my_team(request, teamid, "team_finish")
    raise Http404("Not found")


@login_required
def team_admin(request, teamid=""):
    team_form_admin = TeamFormAdmin(request.POST or None)
    if request.method == "POST" and team_form_admin.is_valid():
        if request.user.is_superuser:
            team = team_form_admin.save(request.user)
            if not team:
                Http404("Wrong values")
            response_data = {}
            response_data["success"] = "true"
            return JsonResponse(response_data)
    raise Http404("Wrong values")


class NewPaymentView(View):
    def post(self, request):
        paymentid = request.POST.get("paymentid", "")
        try:
            team = Team.objects.get(paymentid=paymentid, year=2023)
        except Team.DoesNotExist:
            raise Http404("Team not found")

        payment_method = request.POST.get("payment_method", "visamc")

        cost_now = PaymentsYa.get_cost()
        cost = (team.ucount - team.paid_people) * cost_now

        payment = Payment.objects.create(
            owner=request.user if request.user.is_authenticated else None,
            team=team,
            payment_method=payment_method,
            payment_amount=cost,
            payment_with_discount=cost,
            cost_per_person=cost_now,
            paid_for=team.ucount - team.paid_people,
            additional_charge=team.additional_charge,
            status="draft",
        )
        response_data = {
            "success": "true",
            "payment_id": str(payment.id),
            "sum": cost + team.additional_charge,
            "paymentmethod": payment_method,
        }
        if payment_method == "visamc":
            response_data["yandexwallet"] = settings.YANDEX_WALLET

        if payment_method == "yandexmoney":
            response_data["yandexwallet"] = settings.YANDEX_WALLET

        if payment_method == "sberbank":
            response_data["cardholder_phone"] = settings.SBERBANK_INFO["phone"]
            response_data["cardholder_name"] = settings.SBERBANK_INFO["name"]
            response_data["today_date"] = strftime("%d.%m.%Y", gmtime())

        if payment_method == "sbp":
            response_data["paymentmethod"] = "sbp"
            response_data["cardholder_phone"] = settings.SBP_INFO["phone"]
            response_data["cardholder_name"] = settings.SBP_INFO["name"]
            response_data["today_date"] = strftime("%d.%m.%Y", gmtime())

        PaymentLog.objects.create(
            team=team, payment_method=payment_method, paid_sum=response_data["sum"]
        )

        return JsonResponse(response_data)


class ConfirmPaymentView(View):
    def post(self, request, pk):
        if not request.user.is_superuser:
            raise Http404("Not found")

        balance = Decimal(request.POST.get("balance", 0))

        payment = Payment.objects.get(pk=pk)
        if payment.status == "done":
            return self.update_only_balance(payment, balance)

        payment.status = "done"
        payment.balance = balance
        payment.order = pk

        team = payment.team
        team.paid_people += payment.paid_for
        team.paid_sum += payment.payment_amount

        team.save(update_fields=["paid_people", "paid_sum"])
        payment.save(update_fields=["status", "balance", "order"])
        return HttpResponseRedirect("/payments?status=draft_with_info")

    def update_only_balance(self, payment, balance):
        payment.balance = balance
        payment.save(update_fields=["balance"])
        return HttpResponseRedirect(
            f"/payments?status=done&method=sbp#order{payment.pk}"
        )


class CancelPaymentView(View):
    def post(self, request, pk):
        if not request.user.is_superuser:
            raise Http404("Not found")

        payment = Payment.objects.get(pk=pk)
        payment.status = "cancel"
        payment.save(update_fields=["status"])

        return HttpResponseRedirect("/payments?status=draft_with_info")


class PaymentUp(View):
    def post(self, request, pk):
        if not request.user.is_superuser:
            raise Http404("Not found")

        payment = Payment.objects.get(pk=pk)
        next_payment = (
            Payment.objects.filter(
                order__gt=payment.order,
                status=Payment.STATUS_DONE,
                payment_method__in=["sberbank", "sbp"],
            )
            .order_by("order")
            .first()
        )

        payment.order, next_payment.order = next_payment.order, payment.order
        payment.save(update_fields=["order"])
        next_payment.save(update_fields=["order"])

        return HttpResponseRedirect(
            f"/payments?status=done&method=sbp#order{payment.pk}"
        )


class PaymentDown(View):
    def post(self, request, pk):
        if not request.user.is_superuser:
            raise Http404("Not found")

        payment = Payment.objects.get(pk=pk)
        prev_payment = (
            Payment.objects.filter(
                order__lt=payment.order,
                status=Payment.STATUS_DONE,
                payment_method__in=["sberbank", "sbp"],
            )
            .order_by("-order")
            .first()
        )

        payment.order, prev_payment.order = prev_payment.order, payment.order
        payment.save(update_fields=["order"])
        prev_payment.save(update_fields=["order"])

        return HttpResponseRedirect(
            f"/payments?status=done&method=sbp#order{payment.pk}"
        )


def paymentinfo(request):
    if request.method == "POST":
        new_payment_id = request.POST["paymentid"]
        payment = Payment.objects.filter(id=new_payment_id)[:1]
        if payment:
            payment = payment.get()
            payment.status = "draft_with_info"
            payment.sender_card_number = (
                request.POST["sender_card_number"] + " " + request.POST["payment_sum"]
            )
            pdate = datetime.strptime(request.POST["payment_date"], "%d.%m.%Y")
            payment.payment_date = pdate
            payment.save()

            return JsonResponse(
                {
                    "paymentmethod": payment.payment_method,
                    "success": "true",
                }
            )
    raise Http404("Wrong values")


def get_cost(request):
    return JsonResponse(
        {
            "success": "true",
            "cost": PaymentsYa.get_cost(),
        }
    )


@login_required
def new_team(request):
    if request.method == "POST":
        team = Team()
        team.new_team(request.user, "12h", 4)
        team.save()
        return HttpResponseRedirect("/team/%s" % team.paymentid)


@csrf_exempt
def yandex_payment(request):
    if request.method == "POST":
        payment = PaymentsYa()
        if payment.new_payment(request.POST):
            # send_success_email(
            #     payment.label, payment.withdraw_amount, notification_type
            # )
            return HttpResponse("Ok")
        else:
            raise Http404("Wrong values")
    raise Http404("File not found.")


def import_categories(request):
    if request.user.is_superuser:
        count = import_category_from_sheet()
        return HttpResponse("Updated: %s" % count)
    else:
        raise Http404("File not found.")


def export_payments(request):
    if request.user.is_superuser:
        count = export_payments_to_sheet()
        return HttpResponse("Updated: %s" % count)
    else:
        raise Http404("File not found.")


def sync_googledocs(request):
    if request.user.is_superuser:
        form = Export2GoogleDocsForm(request.POST or None)
        if request.method == "POST" and form.is_valid():
            if form.cleaned_data["sync_type"] == "export_team":
                if export_teams(form.googlekey):
                    return render(
                        request,
                        "website/sync_googledocs.html",
                        {"success": "export", "form": form},
                    )
                return HttpResponse("Export failed")
            elif form.cleaned_data["sync_type"] == "export_team_pretty":
                if export_teams_pretty(form.googlekey):
                    return render(
                        request,
                        "website/sync_googledocs.html",
                        {"success": "export", "form": form},
                    )
                return HttpResponse("Export failed")
            elif form.cleaned_data["sync_type"] == "import_team_numbers":
                count = import_start_numbers_from_sheet(form.googlekey)
                return render(
                    request,
                    "website/sync_googledocs.html",
                    {"success": "import", "count": count, "form": form},
                )
            else:
                raise Http404("File not found.")
        return render(request, "website/sync_googledocs.html", {"form": form})
    else:
        raise Http404("File not found.")


def update_protocol(request):
    if not request.user.is_staff:
        raise Http404("File not found.")

    wb = load_workbook(filename=settings.PROTOCOL_DIR + "protokol.xlsx")
    # grab the active worksheet
    sheet_tabs = {
        "6h": "6ч",
        "12h_ww": "12ч_ЖЖ",
        "12h_mw": "12ч_МЖ",
        "12h_mm": "12ч_ММ",
        "12h_team": "12ч_группа",
        "24h": "24ч",
    }
    distance_max_time = {
        "6h": timedelta(hours=6),
        "12h_ww": timedelta(hours=12),
        "12h_mw": timedelta(hours=12),
        "12h_mm": timedelta(hours=12),
        "12h_team": timedelta(hours=12),
        "24h": timedelta(hours=24),
    }

    for tab in sheet_tabs:
        ws = wb[sheet_tabs[tab]]
        # export KP
        cpoints = ControlPoint.objects.filter(year="2023").order_by("iterator")
        column = 14
        point_column = {}
        for p in cpoints:
            ws.cell(8, column, p.number)
            ws.cell(9, column, p.cost)
            point_column[p.number] = column
            column += 1
        # export teams
        Team().update_places()
        teams = Team.objects.filter(category=tab, year="2023").order_by(
            "place", "start_number"
        )
        teams = [team for team in teams if team.paid_sum > 0]
        line = 10
        for team in teams:
            row = str(line)
            ws["B" + row] = team.start_number
            ws["C" + row] = team.teamname
            athlets = [
                team.athlet1,
                team.athlet2,
                team.athlet3,
                team.athlet4,
                team.athlet5,
                team.athlet6,
            ]
            athlets = ", ".join(athlets[: int(team.paid_people)])
            ws["D" + row] = athlets
            ws["E" + row] = (
                team.start_time + timedelta(hours=5) if team.start_time else ""
            )
            ws["F" + row] = (
                team.finish_time + timedelta(hours=5) if team.finish_time else ""
            )
            if team.distance_time:
                ws["G" + row] = team.distance_time
                if team.distance_time > distance_max_time[tab]:
                    ws["H" + row] = team.distance_time - distance_max_time[tab]
                else:
                    ws["H" + row] = ""
            points = TakenKP.objects.filter(team=team)
            points_sum = 0
            points_count = 0
            for point in points:
                points_sum += point.point.cost
                points_count += 1
                ws.cell(line, point_column[point.point.number], 1)
            ws["I" + row] = points_count
            ws["J" + row] = team.points_sum + team.penalty
            ws["K" + row] = team.penalty
            ws["L" + row] = team.points_sum
            ws["M" + row] = team.place if team.place != 10000 else 0
            if team.dnf:
                ws["M" + row] = "СН"
                ws["K" + row] = "снятие"
            line += 1

    # Save the file
    filename = "protokol2023.xlsx"
    wb.save(settings.PROTOCOL_DIR + filename)
    return render(
        request,
        "website/save_protokol.html",
        {"success": "save", "file_url": settings.PROTOCOL_URL + filename},
    )


def upload_protocol(request):
    if not request.user.is_staff:
        raise Http404("File not found.")

    if request.method == "POST" and request.FILES["myfile"]:
        myfile = request.FILES["myfile"]
        fs = FileSystemStorage()

        curr_time = datetime.now(timezone.utc) + timedelta(hours=5)
        file_prefix = (
            "uploads/"
            + str(curr_time.year)
            + str(curr_time.month).zfill(2)
            + str(curr_time.day).zfill(2)
            + "_"
            + str(curr_time.hour).zfill(2)
            + str(curr_time.minute).zfill(2)
            + str(curr_time.second).zfill(2)
            + "_"
        )

        filename = fs.save(settings.PROTOCOL_DIR + file_prefix + myfile.name, myfile)

        uploaded_file_url = fs.url(settings.PROTOCOL_URL + file_prefix + myfile.name)

        # read this file:
        err, msg = import_file_xlsx(filename)

        return render(
            request,
            "website/simple_upload.html",
            {"uploaded_file_url": uploaded_file_url, "err": err, "msg": msg},
        )
    return render(request, "website/simple_upload.html")


def regulations(request):
    return render(request, "website/regulations.html")


# API _________________________________________________________________
class RaceView(View):
    def get(self, request):
        races = Race.objects.filter(is_active=True)
        data = []
        for race in races:
            data.append(
                {
                    "id": race.id,
                    "name": race.name,
                    "code": race.code,
                    "date": race.date.strftime("%Y-%m-%d"),
                    "is_active": race.is_active,
                }
            )
        return JsonResponse(data, safe=False)


@method_decorator(csrf_exempt, name="dispatch")
class TeamsTimesView(View):
    def post(self, request):
        try:
            times = json.loads(request.body)

            for time_ in times:
                try:
                    team = Team.objects.get(id=time_.get("team_id"))
                except Team.DoesNotExist:
                    continue

                fields = []
                if time_.get("start_time") and not team.start_time:
                    team.start_time = time_.get("start_time")
                    fields.append("start_time")

                if time_.get("finish_time") and not team.finish_time:
                    team.finish_time = time_.get("finish_time")
                    fields.append("finish_time")

                if fields:
                    team.save(update_fields=fields)

            return JsonResponse(
                {"message": f"Teams times updated."},
                status=200,
            )

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data."}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)


@method_decorator(csrf_exempt, name="dispatch")
class PointTagsView(View):
    def post(self, request, race_id):
        try:
            data = json.loads(request.body)

            point_number = data.get("point_number")
            point = self.get_point_by_number(point_number)
            if not point:
                return JsonResponse(
                    {"error": f"Point with number {point_number} not found."},
                    status=404,
                )

            tag_id = data.get("tag_id")
            if tag_id is None:
                return JsonResponse(
                    {"error": "tag_id is a required field."}, status=400
                )

            _, created = PointTag.objects.update_or_create(point=point, tag_id=tag_id)
            if created:
                return JsonResponse(
                    {"message": "PointTag created successfully."}, status=201
                )
            return JsonResponse(
                {"message": f"PointTag with tag_id {tag_id} updated."},
                status=200,
            )

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data."}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    def get_point_by_number(self, point_number):
        try:
            return ControlPoint.objects.get(number=point_number, year=2023)
        except ControlPoint.DoesNotExist:
            return None


def points(request):
    """Возвращает список контрольных пунктов"""
    control_points = (
        ControlPoint.objects.filter(year=2023)
        .order_by("number")
        .values("id", "number", "description", "cost")
    )
    points_ids = [point["id"] for point in control_points]

    tags = PointTag.objects.filter(point_id__in=points_ids).values("point_id", "tag_id")
    tags_by_point = defaultdict(list)
    for tag in tags:
        tags_by_point[tag["point_id"]].append(tag["tag_id"])

    for point in control_points:
        point["tags"] = tags_by_point[point["id"]]

    return JsonResponse(list(control_points), safe=False)


def teams_api(request):
    """Возвращает список команд"""
    query_params = request.GET.get("category", "")
    teams = (
        Team.objects.filter(year=2023, paid_people__gt=0)
        .order_by("id")
        .values(
            "id",
            "paid_people",
            "dist",
            "category",
            "teamname",
            "start_number",
            "teamname",
            "city",
            "organization",
            "year",
            "start_number",
            "start_time",
            "finish_time",
            "place",
            "dnf",
            "penalty",
            "athlet1",
            "athlet2",
            "athlet3",
            "athlet4",
            "athlet5",
            "athlet6",
        )
    )

    if query_params:
        teams = teams.filter(category=query_params)

    teams = list(teams)
    for team in teams:
        paid_people = team.get("paid_people")
        members = [team["athlet1"]]
        if paid_people > 1 and team["athlet2"]:
            members.append(team["athlet2"])

        if paid_people > 2 and team["athlet3"]:
            members.append(team["athlet3"])

        if paid_people > 3 and team["athlet4"]:
            members.append(team["athlet4"])

        if paid_people > 4 and team["athlet5"]:
            members.append(team["athlet5"])

        if paid_people > 5 and team["athlet6"]:
            members.append(team["athlet6"])
        for i in range(6):
            key = f"athlet{i+1}"
            team.pop(key)
        teamname_description = ", ".join(members)
        team["teamname"] = team["teamname"] + f" ({teamname_description})"

    return JsonResponse(list(teams), safe=False)


@csrf_exempt
def upload_photo(request):
    """save file from post request"""
    if request.method != "POST":
        raise Http404("File not found.")

    file = request.FILES.get("photo")
    team_id = request.POST["team_id"]
    point_number = request.POST["point_number"]
    timestamp = request.POST.get("timestamp")
    nfc = request.POST.get("nfc")
    phone_uuid = request.POST.get("phone_uuid")

    team = Team.objects.filter(id=team_id).first()
    if not team:
        return JsonResponse({"error": "team not found"}, status=404)

    uploaded_file_url = ""
    if file:
        start_number = team.start_number
        folder_name = (
            "photos/"
            + f"{start_number}-{team_id}/{phone_uuid}/{point_number}-{file.name}"
        )
        fs = FileSystemStorage()
        filename = fs.save(folder_name, file)
        uploaded_file_url = fs.url(filename)

    # save to db
    TakenKP.objects.create(
        team=team,
        point_number=point_number,
        image_url=uploaded_file_url,
        timestamp=timestamp,
        nfc=nfc,
        phone_uuid=phone_uuid,
    )
    return JsonResponse({"success": True})


class AllTeamsView(View):
    def get(self, request, race_id):
        try:
            race = Race.objects.annotate(
                people_count=Sum("category__team__paid_people"),
            ).get(id=race_id)
        except Race.DoesNotExist:
            # page not found
            raise Http404("File not found.")

        categories = (
            Category.active_objects.filter(race_id=race_id)
            .order_by("order", "id")
            .annotate(
                team_count=Subquery(
                    Team.objects.filter(category2=OuterRef("id"), paid_people__gt=0)
                    .values("category2")
                    .annotate(count=Count("id"))
                    .values("count")[:1]
                )
            )
        )
        teams_ = (
            Team.objects.filter(category2__race_id=race_id, paid_people__gt=0)
            .select_related("category2")
            .order_by(
                "category2__order",
                "start_number",
                "id",
            )
        )
        context = {
            "race": race,
            "categories": categories,
            "teams": teams_,
            "show_category": True,
        }
        return render(request, "teams.html", context)


class AllTeamsResultView(View):
    def get(self, request, race_id):
        try:
            race = Race.objects.annotate(
                people_count=Sum("category__team__paid_people"),
            ).get(id=race_id)
        except Race.DoesNotExist:
            # page not found
            raise Http404("File not found.")

        teams_ = (
            Team.objects.filter(category2__race_id=race_id, paid_people__gt=0)
            .select_related("category2")
            .order_by(
                "category2__order",
                "start_number",
                "id",
            )
        )

        points = ControlPoint.objects.filter(year=2023, cost__gte=0)
        cost = {}
        for p in points:
            cost[p.number] = p.cost

        for team in teams_:
            # members
            members = [team.athlet1]
            if team.paid_people > 1 and team.athlet2:
                members.append(team.athlet2)

            if team.paid_people > 2 and team.athlet3:
                members.append(team.athlet3)
            if team.paid_people > 3 and team.athlet4:
                members.append(team.athlet4)
            if team.paid_people > 4 and team.athlet5:
                members.append(team.athlet5)
            if team.paid_people > 5 and team.athlet6:
                members.append(team.athlet6)
            team.members = ", ".join(members)

            team.points_nfc = []
            summ_nfc = 0
            count_nfc = 0
            nfc_points = (
                TakenKP.objects.filter(team=team.id, timestamp__gt=1697223600000)
                .exclude(nfc="")
                .distinct("point_number")
                .order_by("point_number")
            )
            for p in nfc_points:
                if cost[p.point_number]:
                    summ_nfc += cost[p.point_number]
                    count_nfc += 1
                    team.points_nfc.append(p.point_number)
            team.summ_nfc = summ_nfc
            team.count_nfc = count_nfc

            # PHOTO
            team.points_photo = []
            summ_photo = 0
            count_photo = 0
            photo_points = (
                TakenKP.objects.filter(team=team.id, timestamp__gt=1697223600000)
                .exclude(image_url="")
                .distinct("point_number")
                .order_by("point_number")
            )
            for p in photo_points:
                if cost[p.point_number] and p.point_number not in team.points_nfc:
                    summ_photo += cost[p.point_number]
                    count_photo += 1
                    team.points_photo.append(p.point_number)

            team.summ_photo = summ_photo
            team.count_photo = count_photo

            team.summ_both = team.summ_photo + team.summ_nfc

            team.time_diff = team.finish_time - team.start_time
            if team.finish_time == 0:
                team.time_diff = 0
            seconds = int(team.time_diff / 1000)
            minutes = int(seconds / 60)
            hour = int(minutes / 60)
            team.time = f"{hour}:{minutes%60:02d}:{seconds%60:02d}"

            team.points_nfc = ", ".join(str(p) for p in team.points_nfc)
            team.points_photo = ", ".join(str(p) for p in team.points_photo)

            team.summ_after_penalty = team.summ_both - team.penalty

        teams_ = sorted(
            teams_, key=lambda x: (x.place, x.category, -x.summ_both, x.time_diff)
        )

        context = {
            "race": race,
            "teams": teams_,
            "show_category": True,
        }
        return render(request, "teams_result.html", context)


class TeamPointsView(View):
    def get(self, request, team_id):
        team = Team.objects.filter(id=team_id).first()
        photo_points = (
            TakenKP.objects.filter(team=team_id, timestamp__gt=1697223600000)
            .distinct("image_url")
            .order_by("image_url")
        )
        seconds = int((team.finish_time - team.start_time) / 1000)
        minutes = int(seconds / 60)
        hour = int(minutes / 60)
        time = (f"{hour}:{minutes%60:02d}:{seconds%60:02d}",)

        context = {"team": team, "photo_points": photo_points, "time": time}
        return render(request, "team_points.html", context)


class TeamsView(View):
    def get(self, request, race_id, category_id, *args, **kwargs):
        try:
            race = Race.objects.annotate(
                people_count=Sum("category__team__paid_people")
            ).get(id=race_id)
            category = Category.active_objects.get(race_id=race_id, id=category_id)
        except (Race.DoesNotExist, Category.DoesNotExist):
            # page not found
            raise Http404("File not found.")
        categories = (
            Category.active_objects.filter(race_id=race_id)
            .order_by("order", "id")
            .annotate(
                team_count=Subquery(
                    Team.objects.filter(category2=OuterRef("id"), paid_people__gt=0)
                    .values("category2")
                    .annotate(count=Count("id"))
                    .values("count")[:1]
                )
            )
        )
        teams_ = Team.objects.filter(category2=category, paid_people__gt=0).order_by(
            "start_number",
            "id",
        )
        points = ControlPoint.objects.filter(year=2023, cost__gte=0)
        cost = {}
        for p in points:
            cost[p.number] = p.cost

        for team in teams_:
            summ = 0
            count = 0
            take_points = (
                TakenKP.objects.filter(team=team.id)
                .exclude(nfc="")
                .distinct("point_number")
                .order_by("point_number")
            )
            for p in take_points:
                summ += cost[p.point_number]
                count += 1

            team.summ = summ
            team.count = count

        teams_ = sorted(teams_, key=lambda x: -x.summ)

        context = {
            "race": race,
            "category": category,
            "categories": categories,
            "teams": teams_,
            "show_category": False,
        }
        return render(request, "teams.html", context)


class TeamsViewCsv(View):
    def get(self, request, race_id, category_id, *args, **kwargs):
        teams_ = Team.objects.filter(
            category="6h", paid_people__gt=0, year=2023
        ).order_by(
            "start_number",
            "id",
        )
        points = ControlPoint.objects.filter(year=2023, cost__gte=0)
        cost = {}
        for p in points:
            cost[p.number] = p.cost

        for team in teams_:
            team.points_nfc = []
            summ_nfc = 0
            count_nfc = 0
            nfc_points = (
                TakenKP.objects.filter(team=team.id, timestamp__gt=1697223600000)
                .exclude(nfc="")
                .distinct("point_number")
                .order_by("point_number")
            )
            for p in nfc_points:
                if cost[p.point_number]:
                    summ_nfc += cost[p.point_number]
                    count_nfc += 1
                    team.points_nfc.append(p.point_number)
            team.summ_nfc = summ_nfc
            team.count_nfc = count_nfc

            # PHOTO
            team.points_photo = []
            summ_photo = 0
            count_photo = 0
            photo_points = (
                TakenKP.objects.filter(team=team.id, timestamp__gt=1697223600000)
                .exclude(image_url="")
                .distinct("point_number")
                .order_by("point_number")
            )
            for p in photo_points:
                if cost[p.point_number] and p.point_number not in team.points_nfc:
                    summ_photo += cost[p.point_number]
                    count_photo += 1
                    team.points_photo.append(p.point_number)

            team.summ_photo = summ_photo
            team.count_photo = count_photo

            # BOTH
            team.points_both = []
            summ_both = 0
            count_both = 0
            points_both = (
                TakenKP.objects.filter(team=team.id, timestamp__gt=1697223600000)
                .distinct("point_number")
                .order_by("point_number")
            )
            for p in points_both:
                if cost[p.point_number]:
                    summ_both += cost[p.point_number]
                    count_both += 1
                    team.points_both.append(p.point_number)

            team.summ_both = summ_both
            team.count_both = count_both

        teams_ = sorted(teams_, key=lambda x: -x.summ_both)

        teams_csv = []
        for team in teams_:
            seconds = int((team.finish_time - team.start_time) / 1000)
            minutes = int(seconds / 60)
            hour = int(minutes / 60)
            team_csv = [
                str(team.start_number),
                str(team.teamname),
                str(team.points_nfc),
                str(team.points_photo),
                str(team.count_nfc),
                str(team.count_photo),
                str(team.summ_nfc),
                str(team.summ_photo),
                str(team.summ_both),
                f"{hour}:{minutes%60:02d}:{seconds%60:02d}",
            ]
            team_str = ";".join(team_csv)
            teams_csv.append(team_str)

        result = "<br/>".join(teams_csv)
        return HttpResponse(result)


@user_passes_test(is_admin)
def payment_list(request):
    # select related team to avoid additional queries
    payments = (
        Payment.objects.select_related("team", "team__owner")
        .filter(team__year=2023)
        .order_by("-order", "-id")
    )

    status_filter = request.GET.get("status")
    if status_filter:
        payments = payments.filter(status=status_filter)

    method_filter = request.GET.get("method")
    if method_filter in ("sbp", "sberbank"):
        payments = payments.filter(payment_method__in=("sbp", "sberbank"))
    elif method_filter:
        payments = payments.filter(payment_method=method_filter)

    payments_summ = (
        payments.aggregate(Sum("payment_amount"))["payment_amount__sum"] or 0
    )

    start_sum = 0
    estimate_balances = {}
    for p in payments:
        if p.status == Payment.STATUS_DONE and p.payment_method in ("sbp", "sberbank"):
            estimate_balances[p.id] = start_sum
            start_sum -= p.payment_amount

    for payment_id in estimate_balances:
        estimate_balances[payment_id] -= start_sum

    return render(
        request,
        "payment_list.html",
        {
            "payments": payments,
            "estimate_balances": estimate_balances,
            "payments_summ": round(payments_summ),
            "status": status_filter,
            "method": method_filter,
        },
    )
